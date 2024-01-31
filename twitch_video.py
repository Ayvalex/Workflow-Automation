import pymongo
import cv2
from openpyxl import Workbook
import subprocess
from openpyxl.drawing.image import Image

video_path = 'twitch_nft_demo.mp4'

# Open the video file
video = cv2.VideoCapture(video_path)

# Check if the video file was successfully opened
if not video.isOpened():
    print("Error opening video file")
    exit()

# Get the frame rate of the video
frame_rate = (video.get(cv2.CAP_PROP_FPS))
total_frames = (video.get(cv2.CAP_PROP_FRAME_COUNT))

# Connect to MongoDB
my_client = pymongo.MongoClient("mongodb://localhost:27017/")
my_database = my_client["ProjectTwo"]
my_collection_two = my_database["CollectionTwo"]

# Retrieve the documents containing the frame ranges
work_documents = my_collection_two.find({}, {"Location/Frames to fix": 1})

work_done = []
for doc in work_documents:
    for location_frames in doc["Location/Frames to fix"]:
        frame_range = location_frames[1]  # Access the second element of the list

        # Check if the frame range is specified as a range (e.g., 302-304)
        if "-" in frame_range:
            # Split the frame range into start and end frames
            start_frame, end_frame = map(int, frame_range.split("-"))

            if end_frame <= total_frames:
                work_done.append(location_frames)


def frames_to_timecode(frames, frame_rate):
    # Convert frames to time units
    hours = int(frames // (frame_rate * 3600))
    frames %= frame_rate * 3600
    minutes = int(frames // (frame_rate * 60))
    frames %= frame_rate * 60
    seconds = int(frames // frame_rate)
    frames = int(frames % frame_rate)

    # Format timecode as HH:MM:SS:FF
    timecode = f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"

    return timecode


# Modify work_done list
for i in range(len(work_done)):
    # Check if frame range is given as a single frame or a range
    if "-" in work_done[i][1]:
        # Split the frame range into start and end frames
        start_frame, end_frame = map(int, work_done[i][1].split("-"))

        # Convert frames to timecodes
        start_timecode = frames_to_timecode(start_frame, frame_rate)
        end_timecode = frames_to_timecode(end_frame, frame_rate)

        # Add the timecodes to the list
        work_done[i].append(f"{start_timecode}-{end_timecode}")

# Create a new workbook
workbook = Workbook()

# Get the active sheet
sheet = workbook.active

# Write the headers for each column
sheet['A1'] = 'Location'
sheet['B1'] = 'Frames to fix'
sheet['C1'] = 'Timecode'
sheet['D1'] = 'Thumbnail'

# Write the data to the sheet
for i, row in enumerate(work_done):
    sheet.cell(row=i + 2, column=1, value=row[0])
    sheet.cell(row=i + 2, column=2, value=row[1])
    sheet.cell(row=i + 2, column=3, value=row[2])

    # Extract the middle frame of the frame range
    if "-" in row[1]:
        start_frame, end_frame = map(int, row[1].split("-"))
        middle_frame = (start_frame + end_frame) // 2

        # Generate a unique filename for the extracted frame
        output_filename = f"frame_{i}.jpg"

        # Use FFmpeg to extract the frame
        command = f'ffmpeg -i {video_path} -vf "select=gte(n\,{middle_frame})" -vframes 1 {output_filename}'
        subprocess.call(command, shell=True)

        # Insert the extracted frame as an image in the Excel sheet
        img = Image(output_filename)
        img.width = 96  # Adjust the image width as needed
        img.height = 74  # Adjust the image height as needed
        sheet.column_dimensions['D'].width = 96  # Adjust the column width to accommodate the image
        sheet.row_dimensions[i + 2].height = 74  # Adjust the row height to accommodate the image
        sheet.add_image(img, f'D{i + 2}')  # Insert the image at the corresponding cell

# Save the workbook as an XLS file
workbook.save('new.xls')
